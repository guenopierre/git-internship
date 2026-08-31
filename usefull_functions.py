import pandas as pd 
import numpy as np
import matplotlib.pyplot as plt 
import re
import seaborn
from pathlib import Path
import math
import itertools


from matplotlib.patches import Circle, Polygon
from matplotlib.colors import Normalize
from matplotlib.cm import ScalarMappable
from matplotlib.widgets import Button


from scipy.stats import beta


from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, accuracy_score, precision_score, f1_score
from sklearn.ensemble import RandomForestClassifier



import os
import itertools
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



#%% Flares location 

def convert_prefix_value(s):
    s = str(s)
    prefix = s[0]
    try:
        value = float(s[1:])
    except ValueError:
        value = np.nan
    if prefix == 'X':
        return value * 1e-4
    elif prefix == 'M':
        return value * 1e-5
    elif prefix == 'C':
        return value * 1e-6
    elif prefix == 'B':
        return value * 1e-7
    else:
        return np.nan

def draw_sun_xy(x=None, y=None, figsize=(10, 10), title=None, color='steelblue', ax=None, label = None):
    """
    Trace the solar disk with the Carrington coordinate grid
    and displays points in the form of crosses.
    """
    # ─── Validation ───────────────────────────────────────────────────────────
    if (x is None) != (y is None):
        raise ValueError("The arguments 'x' and 'y' must be provided together.")

    if x is not None and len(x) != len(y):
        raise ValueError(
            f"'x' ({len(x)} éléments) et 'y' ({len(y)} éléments) "
            "doivent avoir la même longueur."
        )

    # ─── Création de la figure SEULEMENT si aucun ax fourni ───────────────────
    if ax is None:
        fig, ax = plt.subplots(figsize=figsize)

    # ─── solar disk ───────────────────────────────────────────────────────
    from matplotlib.patches import Circle
    sun_circle = Circle((0, 0), 1, color='orange', fill=False, linewidth=2.5, zorder=5)
    ax.add_patch(sun_circle)

    # ─── Carrington Grid ────────────────────────────────────────────────────
    lat_lines = np.arange(-75, 90, 15)
    lon_lines = np.arange(0, 360, 30)

    # latitude lines
    for lat in lat_lines:
        lat_rad = np.radians(lat)
        y_grid  = np.sin(lat_rad)
        x_max   = np.sqrt(max(1 - y_grid**2, 0))
        x_arr   = np.linspace(-x_max, x_max, 300)

        lw  = 1.8 if lat == 0 else 1.0
        ls  = '-'  if lat == 0 else '--'
        col = 'red' if lat == 0 else 'gray'

        ax.plot(x_arr, np.full_like(x_arr, y_grid),
                color=col, linewidth=lw, linestyle=ls, alpha=0.6, zorder=2)
        if lat != 0:
            ax.text(-x_max - 0.02, y_grid, f'{lat:+d}°',
                    ha='right', va='center', fontsize=7.5, color='dimgray')

    # longitude lines
    for lon in lon_lines:
        lon_rad = np.radians(lon)
        cos_lon = np.cos(lon_rad)
        sin_lon = np.sin(lon_rad)

        if cos_lon < 0:
            continue

        lat_arr = np.linspace(-np.pi / 2, np.pi / 2, 300)
        x_plot  = -np.cos(lat_arr) * sin_lon
        y_arr   =  np.sin(lat_arr)

        ax.plot(x_plot, y_arr, color='gray', linewidth=1.0,
                linestyle='--', alpha=0.6, zorder=2)

        if abs(sin_lon) <= 1.0:
            ax.text(sin_lon, 0.05, f'{lon}°',
                    ha='center', va='top', fontsize=7.5, color='dimgray')

    # ─── Axis Annotations ─────────────────────────────────────────────────────
    ax.text(-1.15, 0, 'E\n(East)', ha='center', va='center',
            fontsize=11, fontweight='bold', color='steelblue')
    ax.text( 1.15, 0, 'W\n(West)', ha='center', va='center',
            fontsize=11, fontweight='bold', color='steelblue')
    ax.text(0,  1.12, 'N', ha='center', va='bottom',
            fontsize=11, fontweight='bold', color='steelblue')
    ax.text(0, -1.12, 'S', ha='center', va='top',
            fontsize=11, fontweight='bold', color='steelblue')

    # ─── Points (x, y) in the form of crosses ────────────────────────────────────
    if x is not None:
        ax.plot(x, y,
                marker='x',
                markersize=14,
                markeredgewidth=2.5,
                color=color,
                linestyle='none',
                zorder=10,
                label = label)

    # ─── page layout ─────────────────────────────────────────────────────────
    if title:
        ax.set_title(title, fontsize=13, fontweight='bold', pad=15)
        
    ax.set_xlim(-1.35, 1.35)
    ax.set_ylim(-1.30, 1.25)
    ax.set_aspect('equal')
    ax.axis('off')
    
    if label is not None:
        ax.legend(loc='upper right', fontsize=10)

    # ─── Affichage uniquement si on a créé la figure ──────────────────────────
    # (on ne fait ni tight_layout ni show si ax est fourni par l'extérieur)
    return ax

    
def solarcoor2xy(lat, long):
    lat_rad = np.radians(lat)
    lon_rad = np.radians(long)

    # Vérification : le point est-il sur la face visible ?
    # Face visible : lon entre 270° et 360° ou 0° et 90°  (cos(lon) >= 0)
    cos_lon = np.cos(lon_rad)
    visible = cos_lon >= 0

    x =  np.cos(lat_rad) * np.sin(lon_rad)
    y =  np.sin(lat_rad)

    # Inversion de x (Ouest à droite)
    # x_plot = -x
    return x, y, visible 

def format_lat_long_coordinates(df, col='Location'):
    """
    Transforms a string column 'SXXWXX' into columns 'lat' and 'long'.
    
    - S/N: sign of latitude (S = negative, N = positive)
    - W/E: sign of the longitude (W = positive, E = negative) --> solar convention
    - If latitude or longitude is missing, the default value is 0.
    """
    lat_pattern = r'([NS])(\d+\.?\d*)'
    long_pattern = r'([EW])(\d+\.?\d*)'
    
    def parse(s):
        # Latitude
        lat_match = re.search(lat_pattern, s)
        if lat_match:
            lat_sign, lat_val = lat_match.groups()
            lat = float(lat_val) * (1 if lat_sign == 'N' else -1)
        else:
            lat = 0.0
        
        # Longitude
        long_match = re.search(long_pattern, s)
        if long_match:
            long_sign, long_val = long_match.groups()
            long = float(long_val) * (1 if long_sign == 'W' else -1)
        else:
            long = 0.0
        
        return pd.Series([lat, long])
    
    df[['lat', 'long']] = df[col].apply(parse)
    return df


def value_to_color(value, vmin, vmax, cmap=plt.cm.viridis):
    """Convertit une valeur en couleur selon un gradient (cmap)."""
    if vmax == vmin:
        norm_val = 0.5
    else:
        norm_val = (value - vmin) / (vmax - vmin)
    return cmap(norm_val)


def _project(lat_deg, lon_deg):
    """Projection (lat, lon) -> (x, y) sur le disque solaire (vue de face)."""
    lat = np.radians(lat_deg)
    lon = np.radians(lon_deg)
    x = np.cos(lat) * np.sin(lon)
    y = -np.sin(lat)
    return x, y



def draw_sepflare_probability(probability_matrix,
                              grid_resolution=np.arange(-90, 91, 30),
                              figsize=(10, 10),
                              title=None,
                              ax=None,
                              label=None,
                              cmap=plt.cm.viridis,
                              vmax=None):
    """
    Trace le disque solaire, la grille (grid_resolution) et colorie chaque
    case selon la valeur correspondante de probability_matrix (gradient + légende).
    La valeur de chaque case est écrite au centre.

    Si show_uncertainty_table=True, la matrice `uncertainty` est affichée
    sous forme d'un tableau 6x6 sous le plot du soleil.

    probability_matrix : matrice (n-1) x (n-1) avec n = len(grid_resolution).
        - lignes  -> latitude (du haut -90 vers le bas +90)
        - colonnes-> longitude (de gauche -90 vers la droite +90)
    """
    # ─── Validation ───────────────────────────────────────────────────────
    n_cells = len(grid_resolution) - 1
    if probability_matrix.shape != (n_cells, n_cells):
        raise ValueError(
            f"probability_matrix doit être de taille {n_cells}x{n_cells} "
            f"(grid_resolution a {len(grid_resolution)} bornes), "
            f"mais a la forme {probability_matrix.shape}."
        )

    # ─── Figure ───────────────────────────────────────────────────────────
    if ax is None:
        fig, ax = plt.subplots(figsize=figsize)
    else:
        fig = ax.figure

    # ─── Échelle de couleurs ──────────────────────────────────────────────
    finite_vals = probability_matrix[np.isfinite(probability_matrix)]
    vmin = np.nanmin(finite_vals)
    if vmax is None:
        vmax = np.nanmax(finite_vals)
    norm = Normalize(vmin=vmin, vmax=vmax)

    # ─── Coloriage des cases (projetées) ──────────────────────────────────
    n_pts = 30
    for i in range(n_cells):           # ligne = latitude
        lat_low = grid_resolution[i]
        lat_high = grid_resolution[i + 1]
        for j in range(n_cells):       # colonne = longitude
            lon_low = grid_resolution[j]
            lon_high = grid_resolution[j + 1]

            value = probability_matrix[i, j]

            lat_seq = np.linspace(lat_low, lat_high, n_pts)
            lon_seq = np.linspace(lon_low, lon_high, n_pts)

            x1, y1 = _project(np.full(n_pts, lat_low), lon_seq)
            x2, y2 = _project(lat_seq, np.full(n_pts, lon_high))
            x3, y3 = _project(np.full(n_pts, lat_high), lon_seq[::-1])
            x4, y4 = _project(lat_seq[::-1], np.full(n_pts, lon_low))

            xs = np.concatenate([x1, x2, x3, x4])
            ys = np.concatenate([y1, y2, y3, y4])

            verts = np.column_stack([xs, ys])

            if np.isnan(value):
                facecolor = (0.9, 0.9, 0.9, 0.3)
            else:
                facecolor = value_to_color(value, vmin, vmax, cmap)

            poly = Polygon(verts, closed=True, facecolor=facecolor,
                           edgecolor='none', alpha=0.85, zorder=1)
            ax.add_patch(poly)

            lat_c = 0.5 * (lat_low + lat_high)
            lon_c = 0.5 * (lon_low + lon_high)
            xc, yc = _project(lat_c, lon_c)

            if not np.isnan(value):
                r, g, b, _ = facecolor
                luminance = 0.299 * r + 0.587 * g + 0.114 * b
                txt_color = 'white' if luminance < 0.5 else 'black'
                ax.text(xc, yc, f'{value:.1f}',
                        ha='center', va='center',
                        fontsize=8, fontweight='bold',
                        color=txt_color, zorder=6)

    # ─── Disque solaire ───────────────────────────────────────────────────
    sun_circle = Circle((0, 0), 1, color='orange', fill=False,
                        linewidth=2.5, zorder=5)
    ax.add_patch(sun_circle)

    # ─── Grille adaptée à grid_resolution ─────────────────────────────────
    for lat in grid_resolution:
        if abs(lat) >= 90:
            continue
        lat_seq = np.full(200, lat)
        lon_seq = np.linspace(grid_resolution[0], grid_resolution[-1], 200)
        xg, yg = _project(lat_seq, lon_seq)
        col = 'red' if lat == 0 else 'gray'
        lw = 1.8 if lat == 0 else 1.0
        ls = '-' if lat == 0 else '--'
        ax.plot(xg, yg, color=col, linewidth=lw, linestyle=ls,
                alpha=0.6, zorder=3)

    for lon in grid_resolution:
        lat_seq = np.linspace(grid_resolution[0], grid_resolution[-1], 200)
        lon_arr = np.full(200, lon)
        xg, yg = _project(lat_seq, lon_arr)
        ax.plot(xg, yg, color='gray', linewidth=1.0, linestyle='--',
                alpha=0.6, zorder=3)

    # ─── Annotations d'axes ───────────────────────────────────────────────
    ax.text(-1.15, 0, 'E\n(East)', ha='center', va='center',
            fontsize=11, fontweight='bold', color='steelblue')
    ax.text(1.15, 0, 'W\n(West)', ha='center', va='center',
            fontsize=11, fontweight='bold', color='steelblue')
    ax.text(0, 1.12, 'N', ha='center', va='bottom',
            fontsize=11, fontweight='bold', color='steelblue')
    ax.text(0, -1.12, 'S', ha='center', va='top',
            fontsize=11, fontweight='bold', color='steelblue')

    # ─── Légende (colorbar = gradient) ────────────────────────────────────
    sm = ScalarMappable(norm=norm, cmap=cmap)
    sm.set_array([])
    cbar_label = label if label is not None else 'Probability ratio'
    cbar = fig.colorbar(sm, ax=ax, fraction=0.046, pad=0.08, shrink=0.8)
    cbar.set_label(cbar_label, fontsize=11)

    # ─── Mise en page ─────────────────────────────────────────────────────
    if title:
        ax.set_title(title, fontsize=13, fontweight='bold', pad=15)

    ax.set_xlim(-1.35, 1.35)
    ax.set_ylim(-1.30, 1.25)
    ax.set_aspect('equal')
    ax.axis('off')

    return ax





def plot_sep_over_flare_probability(lat_sep, long_sep, lat_flares, long_flares,
                                    label="Probabilty (%)", title="SEP / Flare probability ratio", cmap=plt.cm.Reds, 
                                    grid_resolution=np.arange(-90, 91, 30), vmax=100, ax=None):
    """
    Plot the ratio of Solar Energetic Particle (SEP) events to solar flare events as a probability percentage.

    This function calculates the spatial probability distribution of SEP events relative to solar flare locations
    and visualizes it as a 2D sun heatmap.

    Parameters:
    -----------
    lat_sep : array-like
        Latitude coordinates of SEP events
    long_sep : array-like
        Longitude coordinates of SEP events
    lat_flares : array-like
        Latitude coordinates of solar flare events
    long_flares : array-like
        Longitude coordinates of solar flare events
    label : str, optional
        Label for the colorbar (default: "Probabilty (%)")
    title : str, optional
        Title for the plot (default: "SEP / Flare probability ratio")
    cmap : matplotlib.colors.Colormap, optional
        Colormap to use for the heatmap (default: plt.cm.Reds)
    grid_resolution : array-like, optional
        Grid resolution for the 2D histogram bins (default: np.arange(-90, 91, 30))
    vmax : float, optional
        Maximum value for the color scale (default: 30)
    ax : matplotlib.axes.Axes, optional
        Existing axes to plot on. If None, creates a new figure (default: None)

    Returns:
    --------
    tuple : (ax, probability_matrix, lat_long_sep, lat_long_flares)
        - ax: matplotlib.axes.Axes object containing the plot
        - probability_matrix: 2D array of SEP/flare probability percentages
        - lat_long_sep: 2D histogram of SEP event counts
        - lat_long_flares: 2D histogram of flare event counts
    """

    # Create 2D histograms for SEP and flare coordinates using the specified grid resolution
    lat_long_sep, _, _ = np.histogram2d(
        lat_sep, long_sep, bins=[grid_resolution, grid_resolution]
    )

    # Create 2D histogram for flare coordinates
    lat_long_flares, _, _ = np.histogram2d(
        lat_flares, long_flares, bins=[grid_resolution, grid_resolution]
    )

    # Convert counts to integers for cleaner display
    lat_long_sep = lat_long_sep.astype(int)
    lat_long_flares = lat_long_flares.astype(int)

    # Calculate the probability matrix by dividing SEP counts by flare counts and converting to percentage
    probability_matrix = (lat_long_sep/lat_long_flares)*100
    
    #Calculate the uncertainty (binomiale)
    confidence_intervals, lower_intervals, upper_intervals = binomiale_uncertainty(lat_long_sep, lat_long_flares)
    

    # Call the visualization function to create the heatmap plot
    ax = draw_sepflare_probability(probability_matrix,
                                   title=title,
                                   label=label,
                                   cmap=cmap,
                                   ax=ax,
                                   vmax=vmax)

    # If no axis was provided, show the plot and adjust layout
    if ax is None:
        plt.tight_layout()
        plt.show()

    # Return the plot axis, probability matrix, and both count matrices
    return ax, probability_matrix, lat_long_sep, lat_long_flares




def plot_sep_uncertainty(lat_sep, long_sep, lat_flares, long_flares,
                         grid_resolution=np.arange(-90, 91, 30),
                         figsize=(12, 8),
                         ax=None, 
                         type_uncertainty = 'binomiale'):
    """
    Calcule la matrice de probabilité SEP/flares ainsi que les intervalles
    de confiance binomiaux, puis affiche la matrice confidence_intervals
    (de type [lower, upper]) sous forme d'un tableau.

    Parameters
    ----------
    lat_sep, long_sep : array-like
        Coordonnées (latitude, longitude) des SEP.
    lat_flares, long_flares : array-like
        Coordonnées (latitude, longitude) des flares.
    grid_resolution : array-like
        Bornes de la grille (latitude/longitude).
    figsize : tuple
        Taille de la figure (utilisée seulement si ax est None).
    title : str
        Titre du tableau.
    ax : matplotlib.axes.Axes, optional
        Axe sur lequel dessiner. Si None, un nouvel axe est créé.

    Returns
    -------
    ax : l'axe matplotlib.
    """
    # ─── Histogrammes 2D SEP et flares ────────────────────────────────────
    lat_long_sep, _, _ = np.histogram2d(
        lat_sep, long_sep, bins=[grid_resolution, grid_resolution]
    )

    lat_long_flares, _, _ = np.histogram2d(
        lat_flares, long_flares, bins=[grid_resolution, grid_resolution]
    )

    # ─── Matrice de probabilité (en %) ────────────────────────────────────
    with np.errstate(divide='ignore', invalid='ignore'):
        probability_matrix = (lat_long_sep / lat_long_flares) * 100

    # ─── Intervalles de confiance (binomiale) ─────────────────────────────
    if type_uncertainty == 'binomiale': 
        confidence_intervals, lower_intervals, upper_intervals = \
            binomiale_uncertainty(lat_long_sep, lat_long_flares)
    elif type_uncertainty == 'wilson':
        confidence_intervals, lower_intervals, upper_intervals = \
            wilson_uncertainty(lat_long_sep, lat_long_flares)
    
    # ─── Axe ──────────────────────────────────────────────────────────────
    if ax is None:
        fig, ax = plt.subplots(figsize=figsize)
    ax.axis('off')

    # ─── Affichage du tableau ─────────────────────────────────────────────
    n = confidence_intervals.shape[0]

    # Libellés des lignes (latitude) et colonnes (longitude)
    row_labels = [f'[{grid_resolution[i]}, {grid_resolution[i+1]}]'
                  for i in range(n)]
    col_labels = [f'[{grid_resolution[j]}, {grid_resolution[j+1]}]'
                  for j in range(n)]

    # Contenu : affichage direct des chaînes [lower, upper]
    cell_text = [[str(confidence_intervals[i, j]) for j in range(n)]
                 for i in range(n)]

    table = ax.table(cellText=cell_text,
                     rowLabels=row_labels,
                     colLabels=col_labels,
                     cellLoc='center',
                     loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1, 1.6)
    
    for (row, col), cell in table.get_celld().items(): 
        if row == 0 or col == -1:
            cell.set_text_props(fontweight='bold')
    # Légende des axes du tableau
    
    if type_uncertainty == 'binomiale':  
        ax.set_title("Binomiale confidence intervals", fontsize=13, fontweight='bold')
    
    elif type_uncertainty == 'wilson':
        ax.set_title("Wilson confidence intervals", fontsize=13, fontweight='bold')

    return ax

from scipy.stats import norm

#%% Uncertainties

def binomiale_uncertainty(k_matrix, n_matrix, confidence=0.95, decimals=1):
    """
    Returns a 6x6 matrix of strings in the format"[inf, sup]".

    Settings:
    - k_matrix: Number 6x6 matrix of success counts (B with A)
    - n_matrix: Matrix numpy 6x6 of total counts (A)
    - confidence: Confidence level (default 0.95)
    - decimals: Number of decimal places for display (default is 3)

    Returns:
    - Numpy 6x6 matrix of character strings in the format "[inf, sup]
    """
    alpha = 1 - confidence

    lower = beta.ppf(alpha/2, k_matrix, n_matrix - k_matrix + 1)
    lower*=100
    upper = beta.ppf(1 - alpha/2, k_matrix + 1, n_matrix - k_matrix)
    upper*=100

    formatted = np.array([
        [f"[{round(lower[i,j], decimals)}, {round(upper[i,j], decimals)}]"
         for j in range(6)]
        for i in range(6)
    ])

    return formatted, lower, upper

def wilson_uncertainty(successes, trials, confidence=0.95, decimals=1):
    """
    Calcule les intervalles de confiance de Wilson score pour une proportion
    binomiale, élément par élément sur des matrices.

    Parameters
    ----------
    successes : array-like
        Nombre de succès (ex : SEP).
    trials : array-like
        Nombre d'essais (ex : flares).
    confidence : float
        Niveau de confiance (par défaut 0.95).
    decimals : int
        Nombre de décimales pour l'arrondi (en %).

    Returns
    -------
    confidence_intervals : ndarray of str
        Matrice de chaînes "[lower, upper]" (en %).
    lower_intervals : ndarray of float
        Bornes inférieures (en %).
    upper_intervals : ndarray of float
        Bornes supérieures (en %).
    """
    successes = np.asarray(successes, dtype=float)
    trials = np.asarray(trials, dtype=float)

    # Quantile de la loi normale
    z = norm.ppf(1 - (1 - confidence) / 2)

    lower_intervals = np.full(successes.shape, np.nan)
    upper_intervals = np.full(successes.shape, np.nan)

    with np.errstate(divide='ignore', invalid='ignore'):
        p_hat = successes / trials

    # Calcul de Wilson uniquement là où trials > 0
    mask = trials > 0
    n = trials
    p = p_hat

    # Terme central et demi-largeur
    denom = 1 + (z**2) / n
    center = (p + (z**2) / (2 * n)) / denom
    half_width = (z / denom) * np.sqrt(
        (p * (1 - p) / n) + (z**2) / (4 * n**2)
    )

    lower = (center - half_width) * 100
    upper = (center + half_width) * 100

    lower_intervals[mask] = np.clip(lower[mask], 0, 100)
    upper_intervals[mask] = np.clip(upper[mask], 0, 100)

    # Construction de la matrice de chaînes "[lower, upper]"
    confidence_intervals = np.empty(successes.shape, dtype=object)
    for idx in np.ndindex(successes.shape):
        if mask[idx]:
            lo = round(lower_intervals[idx], decimals)
            up = round(upper_intervals[idx], decimals)
            confidence_intervals[idx] = f'[{lo}, {up}]'
        else:
            confidence_intervals[idx] = 'N/A'

    return confidence_intervals, lower_intervals, upper_intervals


#%% Events time

def time_mean(df1, df2, diff_max = 20):
    """
    To have a positive difference, you suppose that df1 happens earlier than df2

    Parameters
    ----------
    df1 : TYPE
        DESCRIPTION.
    df2 : TYPE
        DESCRIPTION.

    Returns
    -------
    mean : TYPE
        DESCRIPTION.

    """
    df1 = pd.to_datetime(df1, format='%Y-%m-%d %H:%M:%S', errors='coerce')
    df2 = pd.to_datetime(df2, format='%Y-%m-%d %H:%M:%S', errors='coerce')
    df3 = pd.DataFrame({
        'df1': df1,
        'df2': df2
    })
    df_clean = df3.dropna(subset=['df1'])
    df_clean = df_clean.dropna(subset=['df2'])
    
    df_clean['difference'] =  df_clean['df2'] - df_clean['df1'] 
    
    limite = pd.Timedelta(days=diff_max) 
    df_clean = df_clean[df_clean['difference'].abs() <= limite]
    
    mean = df_clean['difference'].mean()
    median = df_clean['difference'].median()
    
    return mean, df_clean, median

#%% Correlation

from scipy.stats import pearsonr, spearmanr

def print_corr(df, col_x, col_y, title=None):
    r_p, p_p = pearsonr(df[col_x], df[col_y])
    r_s, p_s = spearmanr(df[col_x], df[col_y])
    print(f"{title}")
    print(f"  Pearson  r = {r_p:+.3f} (p={p_p:.2e})")
    print(f"  Spearman r = {r_s:+.3f} (p={p_s:.2e})")
    print("---------------------------------------------------")
    return r_p, p_p, r_s, p_s

def logistic_func(x, L, k, x0, b):
    return L / (1 + np.exp(-k * (x - x0))) + b


def correlation_matrix(df, columns, method='pearson',
                       plot=True, interactive = True, cr = False,
                       annotations = True, title = 'S',
                       separators = [], bold_param = 'GSEP flag'):
    """
    Generates a correlation matrix (Pearson or Spearman) for the specified columns,
    by automatically ignoring null values (NaN, NaT, etc.).
 
    Args:
        df (pd.DataFrame): DataFrame containing the data.
        columns (list): A list of column names to include in the matrix.
        method (str): 'pearson' or 'spearman'.
        plot (bool): If True, displays the lower triangular matrix with a heatmap.
 
    Returns:
        pd.DataFrame: Correlation matrix as a DataFrame.
    """
    plt.rcParams['font.serif'] = ['Times New Roman']
    plt.rcParams['font.size'] = 15
    n = len(columns)
    n_samples = len(df)
    corr_matrix = pd.DataFrame(np.zeros((n, n)), index=columns, columns=columns)
 
    for i, col_x in enumerate(columns):
        for j, col_y in enumerate(columns):
            if i <= j:  # On ne calcule que la moitié supérieure (la matrice est symétrique)
                # Filtrer les paires de valeurs non nulles
                mask = df[col_x].notna() & df[col_y].notna()
                x = df.loc[mask, col_x]
                y = df.loc[mask, col_y]
 
                if len(x) < 2 or len(y) < 2:
                    r = np.nan
                else:
                    if method == 'pearson':
                        r, _ = pearsonr(x, y)
                    elif method == 'spearman':
                        r, _ = spearmanr(x, y)
                    else:
                        raise ValueError("The method must be 'pearson' or 'spearman'.")
                    r *= 100  # Conversion en pourcentage
 
                corr_matrix.loc[col_x, col_y] = r
                corr_matrix.loc[col_y, col_x] = r  # Symétrie
 
    if plot:
        # Extraire la partie en bas à gauche (exclut la diagonale)
        lower_triangle = corr_matrix.where(np.tril(np.ones(corr_matrix.shape), k=-1).astype(bool))
 
        # Prendre la valeur absolue pour la couleur
        abs_lower_triangle = lower_triangle.abs()
        fig, ax = plt.subplots(figsize=(20, 10))
        im = ax.imshow(
            abs_lower_triangle,
            cmap='Purples',
            aspect='auto',
            alpha=0.6,
            vmin=0, vmax=75
        )
        fig.colorbar(im, ax=ax, label='|Correlation (%)|')
        # Lines delimitation
        for sep in separators:
            idx = columns.index(sep)
            pos = idx + 0.5
            ax.axhline(y=pos, color='black', linewidth=1.2, zorder=3, alpha = 0.5, linestyle = '--')
            ax.axvline(x=pos, color='black', linewidth=1.2, zorder=3, alpha = 0.5, linestyle = '--')
 
        # Annotations with sign
        if annotations:
            for i in range(n):
                for j in range(n):
                    if i > j:
                        value = lower_triangle.iloc[i, j]
                        is_bold = (j == columns.index(bold_param))
                        ax.text(j, i, f"{value:+.0f}",
                                ha='center', va='center', color='black', fontsize=8,
                                fontweight='bold' if is_bold else 'normal')
 
 
        ax.set_xticks(range(n))
        ax.set_xticklabels(columns, rotation=45, ha='right')
        ax.set_yticks(range(n))
        ax.set_yticklabels(columns)
        ax.set_title(f"{title} - {method} method -  {n} parameters - {n_samples} samples ")
        plt.tight_layout()
 
 
        if interactive:
            def on_click(event):
                # Ignore clicks outside the axes
                if event.inaxes != ax:
                    return
                # Round click coordinates to the nearest cell
                j = int(round(event.xdata))
                i = int(round(event.ydata))
                # Only react to valid lower-triangle cells
                if 0 <= i < n and 0 <= j < n and i > j:
                    col_x = columns[j]  # x-axis parameter
                    col_y = columns[i]  # y-axis parameter
                    scatter_parameters(
                        df[col_x], df[col_y],
                        name_p1=col_x,
                        name_p2=col_y,
                        cr = cr,
                    )
 
            fig.canvas.mpl_connect('button_press_event', on_click)

        plt.show()
 
    return corr_matrix
 
 
def scatter_parameters(p1, p2, name_p1='name_p1', name_p2='name_p2', cr=False):
    """
    Plots two scatter plots (linear and log-log) of p2 vs p1, each with a
    polynomial fit, in a NEW figure.
 
    A "Swap X <-> Y" button lets you invert the two axes interactively:
    the scatter points AND the three fits (linear, quadratic, log-log)
    are recomputed and redrawn on the same figure at each click.
 
    Args:
        cr (bool): if True, min-max normalize p1/p2 to [0, 1] using
            nan-aware min/max (a single NaN in the raw data must NOT be
            allowed to poison every value: plain np.min/np.max return NaN
            if the array contains even one NaN, which would turn the whole
            normalized array into NaN).
 
    Returns:
        (a, b, c, d, e): fit coefficients (linear degree-1: a, b ;
        linear degree-2: c, d, e) corresponding to the orientation
        displayed when the figure is closed (i.e. after any swap).
    """
    p1_orig = np.asarray(p1, dtype=float)
    p2_orig = np.asarray(p2, dtype=float)
 
    if cr:
        p1_min, p1_max = np.nanmin(p1_orig), np.nanmax(p1_orig)
        p2_min, p2_max = np.nanmin(p2_orig), np.nanmax(p2_orig)
        p1_orig = (p1_orig - p1_min) / (p1_max - p1_min)
        p2_orig = (p2_orig - p2_min) / (p2_max - p2_min)
 
    # Etat mutable : orientation courante (normale / inversée) + derniers
    # coefficients de fit, mis a jour a chaque (re)trace.
    state = {'swapped': False, 'coeffs': (np.nan,) * 5}
 
    fig, ax = plt.subplots(1, 2, figsize=(20, 10))
    fig.subplots_adjust(bottom=0.2)  # laisse de la place pour le bouton
 
    def _draw(px_raw, py_raw, name_x, name_y):
        ax[0].clear()
        ax[1].clear()
 
        px_lin, py_lin = px_raw, py_raw
 
        # --- Nettoyage pour le fit linéaire
        finite_mask = np.isfinite(px_lin) & np.isfinite(py_lin)
        px_clean = px_lin[finite_mask]
        py_clean = py_lin[finite_mask]
 
        order = np.argsort(px_clean)
        px_sorted = px_clean[order]
        py_sorted = py_clean[order]
 
        # --- Nettoyage pour le fit log-log
        finite_mask_raw = np.isfinite(px_raw) & np.isfinite(py_raw)
        positive_mask = finite_mask_raw & (px_raw > 0) & (py_raw > 0)
        px_pos = px_raw[positive_mask]
        py_pos = py_raw[positive_mask]
        order_log = np.argsort(px_pos)
        px_pos_sorted = px_pos[order_log]
        py_pos_sorted = py_pos[order_log]
 
        # Polyfit degré 1
        a, b = np.polyfit(px_sorted, py_sorted, 1)
        fit_vals = a * px_sorted + b
 
        # Polyfit degré 2
        c, d, e = np.polyfit(px_sorted, py_sorted, 2)
        fit_vals_2 = c * px_sorted**2 + d * px_sorted + e
 
        # Polyfit log-log (si assez de points positifs et variance non nulle)
        a_log, b_log = np.nan, np.nan
        fit_vals_log = None
        if len(px_pos_sorted) > 2 and np.ptp(np.log10(px_pos_sorted)) > 0:
            a_log, b_log = np.polyfit(np.log10(px_pos_sorted), np.log10(py_pos_sorted), 1)
            fit_vals_log = 10 ** (np.polyval([a_log, b_log], np.log10(px_pos_sorted)))
 
        # --- Plot gauche (linéaire)
        ax[0].scatter(px_clean, py_clean, marker='x', color='purple', alpha=0.5)
        ax[0].plot(px_sorted, fit_vals, '-.', color='blue', alpha=0.9,
                   label=f'DEG 1: {a:.2e} x + {b:.2e}')
        ax[0].plot(px_sorted, fit_vals_2, '-.', color='red', alpha=0.9,
                   label=f'DEG 2: {c:.2e} x² + {d:.2e} x + {e:.2e}')
        ax[0].set_xlabel(f'{name_x}')
        ax[0].set_ylabel(f'{name_y}')
        ax[0].legend()
        ax[0].grid()
        ax[0].set_title(f'{len(px_clean)} parameters')
 
        # --- Plot droite (log-log)
        ax[1].scatter(px_pos, py_pos, marker='x', color='purple', alpha=0.5)
        if fit_vals_log is not None:
            ax[1].plot(px_pos_sorted, fit_vals_log, '-', color='green', alpha=0.9,
                       label=f'LOG: $10^{{{b_log:.2f}}} \\cdot x^{{{a_log:.2f}}}$')
            ax[1].legend()
        ax[1].set_xscale('log'); ax[1].set_yscale('log')
        ax[1].set_xlabel(f'log({name_x})')
        ax[1].set_ylabel(f'log({name_y})')
        ax[1].grid(which='both')
        ax[1].set_title(f'{len(px_pos)} parameters')
 
        fig.suptitle(f'{name_y} = f({name_x})', size='xx-large')
 
        state['coeffs'] = (a, b, c, d, e)
        fig.canvas.draw_idle()
 
    def _current_args():
        if state['swapped']:
            return p2_orig, p1_orig, name_p2, name_p1
        return p1_orig, p2_orig, name_p1, name_p2
 
    def on_swap(event):
        state['swapped'] = not state['swapped']
        _draw(*_current_args())
 
    # Tracé initial
    _draw(*_current_args())
 
    # Bouton d'inversion des axes
    button_ax = fig.add_axes([0.45, 0.03, 0.1, 0.05])
    swap_button = Button(button_ax, 'Swap X ↔ Y')
    swap_button.on_clicked(on_swap)
    fig._swap_button = swap_button  # garde une référence (évite le garbage collection)
 
    plt.show()
 
    return state['coeffs']



def run_pca(df, n_components=2, correlation_circle=False, n_angle_bins=8,
            n_theta_bins=3, n_phi_bins=4):
    """
    n_angle_bins:
        Nombre de secteurs angulaires pour le cercle des corrélations en 2D
        (utilisé seulement si n_components == 2).
    n_theta_bins, n_phi_bins:
        Nombre de bandes de latitude / secteurs d'azimut pour la sphère des
        corrélations en 3D (utilisé seulement si n_components == 3).
        Le nombre total de régions vaut n_theta_bins * n_phi_bins
        (12 par défaut = 3 x 4). Chaque région couvre exactement
        4*pi / (n_theta_bins * n_phi_bins) stéradians : découper z (= cos(theta))
        en tranches égales donne des bandes de latitude de même aire sur la
        sphère (propriété d'Archimède / "hatbox theorem"), qu'on subdivise
        ensuite en secteurs d'azimut égaux -> partition à aire (donc angle
        solide) strictement égale.
 
    Returns
    -------
    pca : sklearn.decomposition.PCA
        L'objet PCA fit sur `df` (utile pour faire pca.transform(...) sur de
        nouvelles données, ex. un jeu de test).
    df_pca : np.ndarray, shape (n_samples, n_components)
        Les données projetées sur les composantes principales.
    scaler : sklearn.preprocessing.StandardScaler
        Le scaler fit sur `df` (utile pour faire scaler.transform(...) sur de
        nouvelles données avant de les passer à pca.transform, en gardant
        exactement le même prétraitement que pour `df`).
    """
    scaler = StandardScaler()
    np_scaled = scaler.fit_transform(df)
    df_scaled = pd.DataFrame(np_scaled, columns=df.columns, index=df.index)
    df_scaled_nanfiltered = df_scaled.fillna(0.0)
    pca = PCA(n_components=n_components)
    df_pca = pca.fit_transform(df_scaled_nanfiltered)
 
    if correlation_circle and n_components == 2:
        from matplotlib.patches import Circle
        eucl_dist = []
        ccircle = []
        for i, j in df_scaled_nanfiltered.T.iterrows():
            corr1 = np.corrcoef(j, df_pca[:, 0])[0, 1]
            corr2 = np.corrcoef(j, df_pca[:, 1])[0, 1]
            ccircle.append((corr1, corr2))
            eucl_dist.append(np.sqrt(corr1**2 + corr2**2))
        names = df_scaled_nanfiltered.columns.tolist()
        # --- Angle avec la verticale (axe Y), en degrés, 0° = vers le haut, sens horaire ---
        angles_deg = [np.degrees(np.arctan2(x, y)) % 360 for (x, y) in ccircle]
        # --- Découpage en secteurs discrets ---
        sector_width = 360 / n_angle_bins
        sector_idx = [int(a // sector_width) % n_angle_bins for a in angles_deg]
        # Couleurs distinctes (qualitatives), pas un dégradé
        base_cmap = plt.cm.get_cmap('tab10' if n_angle_bins <= 10 else 'tab20')
        distinct_colors = [base_cmap(k % base_cmap.N) for k in range(n_angle_bins)]
        fig, axs = plt.subplots(figsize=(6, 6))
        arrow_to_text = {}
        legend_handles = []
        for i in range(len(names)):
            arrow_col = distinct_colors[sector_idx[i]]
            arrow = axs.arrow(0, 0,
                              ccircle[i][0],
                              ccircle[i][1],
                              lw=2,
                              length_includes_head=True,
                              color=arrow_col,
                              fc=arrow_col,
                              head_width=0.02,
                              head_length=0.02,
                              picker=5,
                              label=names[i])  # nom du paramètre pour la légende
            legend_handles.append(arrow)
            x, y = ccircle[i]
            length = np.hypot(x, y)
            if length > 0:
                scale = 1.08
                text_x = (x / length) * scale
                text_y = (y / length) * scale
            else:
                text_x, text_y = 0, 0
            ha = 'left' if x >= 0 else 'right'
            va = 'bottom' if y >= 0 else 'top'
            txt = axs.text(text_x, text_y, names[i], ha=ha, va=va, fontsize=10, visible=False)
            arrow_to_text[arrow] = txt
        circle = Circle((0, 0), 1, facecolor='none', edgecolor='k', linewidth=1, alpha=0.5)
        axs.add_patch(circle)
        axs.set_xlim(-1.2, 1.2)
        axs.set_ylim(-1.2, 1.2)
        axs.set_xlabel("Pearson Correlation Coefficient with PC 1")
        axs.set_ylabel("Pearson Correlation Coefficient with PC 2")
        # Légende discrète : une entrée par paramètre, couleur = angle avec la verticale
        axs.legend(handles=legend_handles,
                   loc='center left',
                   bbox_to_anchor=(1.02, 0.5),
                   fontsize=8,
                   title="Parameters\n(color = angle with vertical)",
                   frameon=False)
        def on_pick(event):
            clicked_arrow = event.artist
            if clicked_arrow in arrow_to_text:
                text_label = arrow_to_text[clicked_arrow]
                text_label.set_visible(not text_label.get_visible())
                fig.canvas.draw_idle()
        fig.canvas.mpl_connect('pick_event', on_pick)
        plt.tight_layout()
        plt.show()
 
    elif correlation_circle and n_components == 3:
        # projection '3d' est auto-enregistrée par mpl_toolkits.mplot3d dès l'import
        from mpl_toolkits.mplot3d import Axes3D  # noqa: F401
 
        ccircle = []
        eucl_dist = []
        for i, j in df_scaled_nanfiltered.T.iterrows():
            corr1 = np.corrcoef(j, df_pca[:, 0])[0, 1]
            corr2 = np.corrcoef(j, df_pca[:, 1])[0, 1]
            corr3 = np.corrcoef(j, df_pca[:, 2])[0, 1]
            ccircle.append((corr1, corr2, corr3))
            eucl_dist.append(np.sqrt(corr1**2 + corr2**2 + corr3**2))
        names = df_scaled_nanfiltered.columns.tolist()
 
        # --- Découpage de la sphère en n_theta_bins x n_phi_bins régions de même angle solide ---
        n_regions = n_theta_bins * n_phi_bins
        theta_edges = np.linspace(-1.0, 1.0, n_theta_bins + 1)  # bornes sur z = cos(theta)
        phi_width = 360.0 / n_phi_bins
 
        region_idx = []
        for (x, y, z) in ccircle:
            r = np.sqrt(x**2 + y**2 + z**2)
            if r > 0:
                xn, yn, zn = x / r, y / r, z / r
            else:
                xn, yn, zn = 0.0, 0.0, 1.0
            # bande de latitude (aire égale car bornes équiréparties sur z)
            band = int(np.digitize(zn, theta_edges) - 1)
            band = min(max(band, 0), n_theta_bins - 1)
            # secteur azimutal
            phi_deg = np.degrees(np.arctan2(yn, xn)) % 360
            sector = int(phi_deg // phi_width) % n_phi_bins
            region_idx.append(band * n_phi_bins + sector)
 
        # Couleurs distinctes (qualitatives), même stratégie que le cercle 2D
        base_cmap = plt.cm.get_cmap('tab10' if n_regions <= 10 else 'tab20')
        distinct_colors = [base_cmap(k % base_cmap.N) for k in range(n_regions)]
 
        fig = plt.figure(figsize=(7, 7))
        axs = fig.add_subplot(111, projection='3d')
 
        arrow_to_text = {}
        legend_handles = []
        legend_labels = []
        for i in range(len(names)):
            x, y, z = ccircle[i]
            arrow_col = distinct_colors[region_idx[i]]
            arrow = axs.quiver(0, 0, 0, x, y, z,
                                color=arrow_col,
                                linewidth=2,
                                arrow_length_ratio=0.12,
                                picker=5)
            # Line3DCollection n'a pas un rendu de légende idéal -> proxy 2D de la même couleur
            legend_handles.append(plt.Line2D([0], [0], color=arrow_col, lw=2))
            legend_labels.append(names[i])
 
            length = np.sqrt(x**2 + y**2 + z**2)
            if length > 0:
                scale = 1.08
                text_x, text_y, text_z = (x / length) * scale, (y / length) * scale, (z / length) * scale
            else:
                text_x, text_y, text_z = 0, 0, 0
            txt = axs.text(text_x, text_y, text_z, names[i], fontsize=9, visible=False)
            arrow_to_text[arrow] = txt
 
        # --- Sphère unité de référence, translucide ---
        u = np.linspace(0, 2 * np.pi, 60)
        v = np.linspace(0, np.pi, 30)
        sx = np.outer(np.cos(u), np.sin(v))
        sy = np.outer(np.sin(u), np.sin(v))
        sz = np.outer(np.ones_like(u), np.cos(v))
        axs.plot_surface(sx, sy, sz, color='grey', alpha=0.06, linewidth=0, shade=False)
 
        # --- Lignes de découpage des n_regions zones (bandes de latitude + méridiens) ---
        t = np.linspace(0, 2 * np.pi, 100)
        for edge_z in theta_edges[1:-1]:  # bornes internes seulement (pas les pôles)
            edge_r = np.sqrt(max(0.0, 1 - edge_z**2))
            axs.plot(edge_r * np.cos(t), edge_r * np.sin(t), edge_z,
                     color='k', lw=0.7, alpha=0.4)
        vv = np.linspace(0, np.pi, 60)
        for phi in np.linspace(0, 2 * np.pi, n_phi_bins, endpoint=False):
            axs.plot(np.sin(vv) * np.cos(phi), np.sin(vv) * np.sin(phi), np.cos(vv),
                     color='k', lw=0.7, alpha=0.4)
 
        axs.set_xlim(-1, 1)
        axs.set_ylim(-1, 1)
        axs.set_zlim(-1, 1)
        axs.set_xlabel("Correlation with PC 1")
        axs.set_ylabel("Correlation with PC 2")
        axs.set_zlabel("Correlation with PC 3")
        axs.set_box_aspect([1, 1, 1])
 
        axs.legend(legend_handles, legend_labels,
                   loc='center left',
                   bbox_to_anchor=(1.05, 0.5),
                   fontsize=8,
                   title=f"Parameters\n(color = 1 of {n_regions} equal-solid-angle\nregions, "
                         f"{4 * np.pi / n_regions:.3f} sr each)",
                   frameon=False)
 
        def on_pick(event):
            clicked_arrow = event.artist
            if clicked_arrow in arrow_to_text:
                text_label = arrow_to_text[clicked_arrow]
                text_label.set_visible(not text_label.get_visible())
                fig.canvas.draw_idle()
        fig.canvas.mpl_connect('pick_event', on_pick)
 
        plt.tight_layout()
        plt.show()
 
    elif correlation_circle:
        print(f"correlation_circle plotting is only implemented for n_components in "
              f"{{2, 3}} (got n_components={n_components}); skipping the plot.")
 
    return pca, df_pca, scaler

#%% Flux Timeseries



def plot_SEP_event(sep_dictionary, index_sep=None, date_time_sep=None, 
                   int_channel = False, channels=[1],
                   log_xray = False, log_diff_channels = False, log_int_channel = False, 
                   colors_channels=None, color_xray='red', color_int = 'darkgreen', 
                   save_fig_path=None, ax=None):   
    
    
    # ------------ Understanding the request -------------------------------------
    if index_sep is None and date_time_sep is None:
        print("Insert an SEP event indicator (index or date)")
        return
    elif index_sep is not None and date_time_sep is not None:
        print("One SEP indicator is enough (index or date)")
        return

    if index_sep is not None:
        keys = list(sep_dictionary.keys())
        SEP_event = sep_dictionary[keys[index_sep]]
    else:
        SEP_event = sep_dictionary[f'{date_time_sep}']

    # --- Create the figure only if no ax is provided ---
    if ax is None:
        fig, ax = plt.subplots(figsize=(15, 8))
    else:
        fig = ax.figure
        
    #------------ X-ray flux -------------------------------------------
    ax.plot(SEP_event['xrayl'], color=color_xray, alpha=0.6, label='X-ray flux (GOES) [5\']')

    ax.set_ylim(0, SEP_event['xrayl'].max() * 1.2)
    
    if log_xray == True: 
        ax.set_yscale('log')  
        
    ax.set_ylabel('X-ray flux (W/m²)', color=color_xray)
    ax.tick_params(axis='y', labelcolor=color_xray)
    
    classification_flares = ['C', 'M', 'X']
    
    for threshold, label in zip((1e-6, 1e-5, 1e-4), classification_flares):
        ax.axhline(y=threshold, color=color_xray, linestyle='--', linewidth=0.8, alpha=0.5)
        
        if ax.get_ylim()[0] <= threshold <= ax.get_ylim()[1]:  
            ax.text(
                ax.get_xlim()[0]+0.05,
                threshold,
                label,
                color=color_xray,
                fontsize=7,
                va='bottom',
                ha='left',
                alpha=0.7,
            )

        
    ax.grid(True, which='major', linestyle='-', linewidth=0.5, alpha=0.5)
    ax.grid(True, which='minor', linestyle=':', linewidth=0.3, alpha=0.3)

    
    #------------ Differential Channels -------------------------------------------
    ax2 = ax.twinx()
    
    label_diff_channels = ['5.00-7.23 MeV [5\']', '7.23-10.46 MeV [5\']', '10.46-15.12 MeV [5\']', 
                           '15.12-21.87 MeV [5\']', '21.87-31.62 MeV  [5\']', '45.73-66.13 MeV [5\']',
                           '66.13-95.64 MeV [5\']', '95.64-138.3 MeV [5\']', '138.3-200.0 MeV [5\']',
                           '200.0-289.2 MeV [5\']']
    
    if colors_channels is None:
        colors_channels = [f'C{idx}' for idx in range(len(channels))]
    
    for i, color_channel in zip(channels, colors_channels):
        ax2.plot(SEP_event[f'F{i}'], color=color_channel, label= label_diff_channels[i])
    
     
    smallest_channel = min(channels)
    ref_data = SEP_event[f'F{smallest_channel}']
    scale_max = ref_data.max() * 1.2
    
    ax2.set_ylim(0, scale_max)
    
    if log_diff_channels == True: 
        ax2.set_yscale('log')
        
        
    ax2.set_ylabel(f'Differential channels F{channels} (cm-2.s-1.sr-1.MeV-1)', color=colors_channels[0])
    ax2.tick_params(axis='y', labelcolor=colors_channels[0])
    
    #------------ Integral channel -------------------------------------------
    if int_channel == True:
        fig.subplots_adjust(right=0.80)
        
        ax3 = ax.twinx()
        ax3.spines['right'].set_position(('axes', 1.10))
        ax3.set_frame_on(True)
        ax3.patch.set_visible(False)
        ax3.plot(SEP_event['ZPGT10W_CORR'], color=color_int, alpha = 0.9, label='>10MeV [5\']')
        
        ax3.set_ylim(0, SEP_event['ZPGT10W_CORR'].max()*1.2)
        
        if log_int_channel == True: 
            ax3.set_yscale('log')
        
        ax3.set_ylabel('Integral channel >10 MeV (pfu = cm-2.s-1.sr-1)', color=color_int)
        ax3.tick_params(axis='y', labelcolor=color_int)
        
        ax3.axhline(y=10, color=color_int, linestyle='--', linewidth=0.8, alpha=0.5)
        
        if ax3.get_ylim()[0] <= 10 <= ax3.get_ylim()[1]:  
            ax3.text(
                    ax3.get_xlim()[-1]-0.2,
                    10,
                    'SWPC threshold',              
                    color=color_int,
                    fontsize=10,
                    va='bottom',
                    ha='left',
                    alpha=0.7,
                ) 
           
    
    #--------------------time events------------------------------------------------------
    # Tracé des trois traits verticaux pour les événements temporels
    time_events = []
    
        # ('GSEP_timestamp', 'GSEP timestamp', 'k', '-'),
        # ('cdaw_start_time', 'CDAW start time', 'olive', '-'),
        # ('cdaw_max_time', 'CDAW max time', 'darkgreen', '-'),
        # ('cme_1st_app_time', 'CME 1st app time', 'sienna', '--'), 
        # ('cme_launch_time', 'CME launch time', 'brown', '--'), 
        # ('fl_start_time', 'Flare start time', 'red', '-.'), 
        # ('fl_peak_time', 'Flare peak time', 'darkred', '-.'), 
    
    for key, label_event, color_event, linestyle in time_events:
        value = SEP_event[key].iloc[0] 
       
        if pd.isna(value):
            continue
        
        time_value = pd.to_datetime(value)  
        ax.axvline(x=time_value, color=color_event, linestyle=linestyle, 
                   linewidth=2.5, alpha=0.25, label=label_event)
        
        
    # Récupération des handles et labels APRÈS avoir tracé les axvlines
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    
    if int_channel == True:
        lines3, labels3 = ax3.get_legend_handles_labels()
        ax.legend(lines1 + lines2 + lines3, labels1 + labels2 + labels3, loc='upper left')
        plt.setp(ax3.get_xticklabels(), visible=False) 
    else:
        ax.legend(lines1 + lines2, labels1 + labels2, loc='upper left')

    plt.setp(ax2.get_xticklabels(), visible=False) 

    
    #--------------------titles------------------------------------------------------

    if int_channel == True: 
        ax.set_title(f'X ray flux, F{channels} & >10MeV ')
        
    else:
        ax.set_title(f'X ray flux & F{channels}')
    
    fig.tight_layout()
    
    if save_fig_path is not None:
        fig.savefig(f'{save_fig_path}')

    return ax  


def plot_SEP_event_interactive_keys(sep_dictionary, channels=[1], colors_channels=None, log=False, int_channel=False, 
                                    log_xray = False, log_diff_channels = False, log_int_channel = False, color_int = 'darkgreen',
                                     color_xray='red', start_index=0, max_index=423):
    keys = list(sep_dictionary.keys())
    state = {'index': start_index}
    selected_indexes = []

    fig, ax = plt.subplots(figsize=(15, 8))

    def draw(j):
        ax.clear()

        for extra_ax in fig.axes:
            if extra_ax is not ax:
                extra_ax.remove()

        plot_SEP_event(sep_dictionary, 
                       index_sep=j,
                       int_channel=int_channel,
                       channels=channels,
                       log_xray=log_xray, 
                       log_diff_channels=log_diff_channels, 
                       log_int_channel=log_int_channel,
                       colors_channels=colors_channels, 
                       color_xray=color_xray, 
                       color_int=color_int,
                       ax=ax
                       )

        ax.set_title(
            f"X-ray flux & F{channels} & >10MeV channel  |  Index {j}/{max_index-1}  "
            f"({keys[j]})  -  SPACE: skip  |  ENTER: select"
        )
        fig.tight_layout()
        fig.canvas.draw_idle()

    def next_image():
        state['index'] += 1
        if state['index'] > max_index:
            print("Last index reached. Closed.")
            print("Selected indexes:", selected_indexes)
            plt.close(fig)
            return
        draw(state['index'])

    def on_key(event):
        if event.key == ' ':            
            next_image()
        elif event.key == 'enter':      # 
            selected_indexes.append(state['index'])
            print(f"Index {state['index']} saved ({keys[state['index']]})")
            next_image()

    fig.canvas.mpl_connect('key_press_event', on_key)
    draw(state['index'])
    plt.show()

    return selected_indexes

#%% Flare time density 

def compute_event_counts(flares_time_peak,
                         start='1976-03-14',
                         end='2025-01-31',
                         window_hours=1,    # <-- fenêtre paramétrable
                         step_hours=1):     # <-- pas d'échantillonnage
    """
    Counts the number of events in a preceding sliding window
    each sampling moment.

    Settings
    ----------
    flares_time_peak: pd.Series or array of Timestamps (event dates)
    start, end  : boundaries of the period
    window_hours  : width of calculation window (in hours)
    step_hours  : interval between two evaluated times (in hours)

    Returns
    --------
    DataFrame indexed by (date, hour) with a 'count_events' column
    """

    # --- 1. Préparer et trier les événements (nécessaire pour searchsorted) ---
    events = pd.to_datetime(pd.Series(flares_time_peak)).sort_values()
    events_ns = events.values.astype('datetime64[ns]')
    print("fin étape 1")

    # --- 2. Générer tous les instants d'échantillonnage ---
    sample_times = pd.date_range(start=start,
                                 end=pd.Timestamp(end) + pd.Timedelta(days=1),
                                 freq=f'{step_hours}h',
                                 inclusive='left')
    print("fin étape 2")

    # --- 3. Calcul vectorisé du nombre d'événements par fenêtre ---
    end_window = sample_times.values.astype('datetime64[ns]')
    start_window = (sample_times - pd.Timedelta(hours=window_hours)).values.astype('datetime64[ns]')

    # searchsorted donne la position d'insertion -> nombre d'événements
    # avant chaque borne, en O(log n) au lieu d'un filtrage complet
    idx_end = np.searchsorted(events_ns, end_window, side='left')    # < end_window
    idx_start = np.searchsorted(events_ns, start_window, side='left') # >= start_window

    counts = idx_end - idx_start
    print("fin étape 3")

    # --- 4. Construire le DataFrame résultat ---
    result = pd.DataFrame({
        'datetime': sample_times,
        'count_events': counts
    })
    result['date'] = result['datetime'].dt.normalize()
    result['hour'] = result['datetime'].dt.hour
    result = result.set_index(['date', 'hour'])[['count_events']]
    print("fin étape 4")

    return result


def plot_timeseries(df, resample=None, figsize=(15, 5)):
    """
    Trace count_events en fonction du temps.

    resample : None, 'D', 'W', 'ME', 'YE'... pour agréger
               (utile car 428k points = illisible et lourd)
    """
    # Reconstruire un index datetime à partir de (date, hour)
    s = df['count_events'].copy()
    dt_index = (s.index.get_level_values('date')
                + pd.to_timedelta(s.index.get_level_values('hour'), unit='h'))
    s.index = dt_index
    s = s.sort_index()

    # Agrégation optionnelle
    if resample:
        s = s.resample(resample).sum()

    fig, ax = plt.subplots(figsize=figsize)
    ax.plot(s.index, s.values, lw=0.7, color='steelblue')
    ax.set_xlabel('Date')
    ax.set_ylabel("Nombre d'événements")
    title = "Événements au cours du temps"
    if resample:
        title += f" (agrégé par '{resample}')"
    ax.set_title(title)
    ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.show()


import seaborn as sns

def plot_heatmap(df, figsize=(14, 8), cmap='inferno', log_scale=False):
    """
    Heatmap : axe X = heures, axe Y = jours (ou inversement).
    df doit être indexé par (date, hour).
    """
    # Passer de l'index multi à une matrice date × hour
    pivot = df['count_events'].unstack(level='hour')  # lignes=date, colonnes=hour

    # Optionnel : échelle logarithmique si valeurs très étalées
    data = np.log1p(pivot) if log_scale else pivot

    fig, ax = plt.subplots(figsize=figsize)
    sns.heatmap(
        data,
        cmap=cmap,
        cbar_kws={'label': 'log(count+1)' if log_scale else "Nombre d'événements"},
        ax=ax
    )
    ax.set_xlabel('Heure de la journée')
    ax.set_ylabel('Date')
    ax.set_title('Nombre d\'événements par heure')

    # Alléger l'axe Y (sinon 17855 labels illisibles)
    n_ticks = 15
    yticks = np.linspace(0, len(pivot) - 1, n_ticks, dtype=int)
    ax.set_yticks(yticks + 0.5)
    ax.set_yticklabels([pivot.index[i].strftime('%Y-%m') for i in yticks], rotation=0)

    plt.tight_layout()
    plt.show()
    
    
def plot_timeseries_overlay(df, resamples=('D', 'W', 'ME', 'YE'),
                            figsize=(15, 6), normalize=True):
    """
    Overlays multiple aggregations on a single axis.
    normalize=True -> reports each curve to a comparable daily average.
    """
    s = df['count_events'].copy()
    dt_index = (s.index.get_level_values('date')
                + pd.to_timedelta(s.index.get_level_values('hour'), unit='h'))
    s.index = dt_index
    s = s.sort_index()

    # Nb de jours par période pour ramener à une "moyenne par jour"
    days_per = {'D': 1, 'W': 7, 'ME': 30.44, 'YE': 365.25}
    labels = {'D': 'daily', 'W': 'weekly', 'ME': 'monthly', 'YE': 'annual'}
    colors = {'D': 'lightgray', 'W': 'steelblue', 'ME': 'darkorange', 'YE': 'crimson'}

    fig, ax = plt.subplots(figsize=figsize)

    for r in resamples:
        s_agg = s.resample(r).sum()
        if normalize:
            s_agg = s_agg / days_per[r]      # -> événements/jour moyens
        ax.plot(s_agg.index, s_agg.values,
                lw=1.2, label=labels.get(r, r), color=colors.get(r))

    ax.set_xlabel('Date')
    ax.set_ylabel("events/ day" if normalize else "events")
    ax.set_title("Time density (different resolutions)")
    ax.legend()
    ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.show()

#%% ML

# (keep your existing imports too: RandomForestClassifier, train_test_split,
#  confusion_matrix, plt, sns, run_pca, etc.)



from datetime import datetime
import numpy as np


from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.svm import SVC
from openpyxl import Workbook


def run_ml(
    df: pd.DataFrame, 
    inputs: list, 
    output: str, 
    test_all_combinations: bool = False, 
    save_dir: str = None,
    model: str = 'RandomForest', 
    train_split: float = 0.8, 
    random_state: int = 42
):
    """
    Run machine learning classification on a dataset.
    
    Parameters
    ----------
    df : pd.DataFrame
        The input dataframe containing features and the target.
    inputs : list of str
        List of column names to be used as features.
    output : str
        Column name of the binary target (0 or 1).
    test_all_combinations : bool, default False
        If True, tests all possible combinations of the provided 'inputs' and saves to Excel.
        If False, runs a single model using all 'inputs' and prints results to console.
    save_dir : str, optional
        Directory path to save the Excel file (required if test_all_combinations=True).
    model : str, default 'RandomForest'
        Model to train. Supported: 'RandomForest', 'LogisticRegression', 'GradientBoosting', 'SVC'.
    train_split : float, default 0.8
        Proportion of the dataset to include in the train split.
    random_state : int, default 42
        Random seed for reproducibility.
    """
    
    # ---------------------------------------------------------
    # 0. Input validation
    # ---------------------------------------------------------
    missing_cols = [col for col in inputs + [output] if col not in df.columns]
    if missing_cols:
        raise ValueError(f"The following columns are missing in the DataFrame: {missing_cols}")
        
    if test_all_combinations and save_dir is None:
        raise ValueError("When test_all_combinations is True, you must provide a valid 'save_dir'.")

    supported_models = ['RandomForest', 'LogisticRegression', 'GradientBoosting', 'SVC']
    if model not in supported_models:
        raise ValueError(f"Model '{model}' is not supported. Choose from {supported_models}.")

    test_size = 1.0 - train_split

    # ---------------------------------------------------------
    # Helper Function: Train and Evaluate a specific feature set
    # ---------------------------------------------------------
    def _train_and_evaluate(selected_features):
        X = df[selected_features].copy()
        y = df[output].copy().astype(int)

        # Train/test split with stratification for balanced classes
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=test_size, random_state=random_state, stratify=y
        )

        # Model instantiation
        if model == 'RandomForest':
            clf = RandomForestClassifier(random_state=random_state, class_weight='balanced')
        elif model == 'LogisticRegression':
            clf = LogisticRegression(random_state=random_state, max_iter=1000, class_weight='balanced')
        elif model == 'GradientBoosting':
            clf = GradientBoostingClassifier(random_state=random_state)
        elif model == 'SVC':
            clf = SVC(random_state=random_state, kernel='linear', class_weight='balanced')

        # Run
        clf.fit(X_train, y_train)
        y_pred = clf.predict(X_test)

        # Metrics calculation
        cm = confusion_matrix(y_test, y_pred, labels=[0, 1])
        tn, fp, fn, tp = cm.ravel()

        metrics = {}
        metrics['Accuracy'] = (tp + tn) / (tp + tn + fp + fn) if (tp + tn + fp + fn) > 0 else np.nan
        metrics['POD'] = tp / (tp + fn) if (tp + fn) > 0 else np.nan
        metrics['FAR'] = fp / (tp + fp) if (tp + fp) > 0 else np.nan
        metrics['Precision'] = tp / (tp + fp) if (tp + fp) > 0 else np.nan
        metrics['F1 Score'] = (
            2 * (metrics['Precision'] * metrics['POD']) / (metrics['Precision'] + metrics['POD'])
            if (metrics['Precision'] + metrics['POD']) > 0 else np.nan
        )
        
        pofd = fp / (fp + tn) if (fp + tn) > 0 else np.nan
        metrics['TSS'] = metrics['POD'] - pofd if not (np.isnan(metrics['POD']) or np.isnan(pofd)) else np.nan
        
        hss_denom = (tp + fn) * (fn + tn) + (tp + fp) * (fp + tn)
        metrics['HSS'] = 2 * (tp * tn - fp * fn) / hss_denom if hss_denom > 0 else np.nan

        # Feature importance (if available for the model)
        feature_importances = {}
        if hasattr(clf, "feature_importances_"):
            importances = clf.feature_importances_
            feature_importances = dict(zip(selected_features, importances))
        elif hasattr(clf, "coef_"):
            importances = np.abs(np.ravel(clf.coef_))
            feature_importances = dict(zip(selected_features, importances))

        return cm, metrics, feature_importances

    # ---------------------------------------------------------
    # Logic Branch A: Single Run (Console Output)
    # ---------------------------------------------------------
    if not test_all_combinations:
        cm, metrics, feature_importances = _train_and_evaluate(inputs)
        tn, fp, fn, tp = cm.ravel()
        
        print("=" * 60)
        print(f"RESULTS - Model: '{model}'")
        print("=" * 60)
        print("Confusion Matrix:")
        print("                 Predicted 0     Predicted 1")
        print(f"Actual 0        |      {tn:>6}          {fp:>6}")
        print(f"Actual 1        |      {fn:>6}          {tp:>6}")
        print("-" * 60)
        for m_name, m_val in metrics.items():
            print(f"{m_name:<11}: {m_val:.4f}")
        
        if feature_importances:
            print("\n" + "=" * 60)
            print("FEATURE IMPORTANCE")
            print("=" * 60)
            # Sort dict by value descending
            sorted_fi = sorted(feature_importances.items(), key=lambda item: item[1], reverse=True)
            for feat_name, imp in sorted_fi:
                print(f"{feat_name:<30}: {imp:.4f}")
        return

    # ---------------------------------------------------------
    # Logic Branch B: Loop Combinations (Excel Save Approach 1)
    # ---------------------------------------------------------
    os.makedirs(save_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(save_dir, f"results_{date_str}.xlsx")
    
    METRIC_LABELS = ['Accuracy', 'POD', 'FAR', 'Precision', 'F1 Score', 'TSS', 'HSS']
    
    # 1. Initialize Excel Workbook and base layout
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    row_of = {}
    r = 1
    
    # Build column A architecture (like old run_ml_sep)
    ws.cell(row=r, column=1, value='Configuration'); row_of['Configuration'] = r; r += 1
    for p in inputs:
        ws.cell(row=r, column=1, value=p); row_of[p] = r; r += 1
    ws.cell(row=r, column=1, value='Output'); row_of['Output'] = r; r += 1
    
    for m in METRIC_LABELS:
        ws.cell(row=r, column=1, value=m); row_of[m] = r; r += 1
        
    ws.cell(row=r, column=1, value='Feature Importance'); row_of['Feature Importance'] = r; r += 1
    for p in inputs:
        fi_label = f'FI: {p}'
        ws.cell(row=r, column=1, value=fi_label); row_of[fi_label] = r; r += 1
        
    wb.save(filepath)

    # 2. Generate all combinations
    num_features = len(inputs)
    combos = []
    for size in range(1, num_features + 1):
        combos.extend(itertools.combinations(inputs, size))
    
    total = len(combos)
    print(f"Starting execution for {total} combinations...")
    print(f"Results will be appended live to: {filepath}")

    # 3. Iterate through combinations and append columns
    for i, combo in enumerate(combos, start=1):
        combo_list = list(combo)
        print(f"Running [{i}/{total}] : {len(combo_list)} feature(s)...", end="\r")
        
        try:
            _, metrics, fi = _train_and_evaluate(combo_list)
        except Exception as e:
            print(f"\nERROR on combination {combo_list}: {e}")
            continue
            
        # Write results to Excel
        new_col = ws.max_column + 1
        
        config_desc = f"Model: {model} | TrainSplit: {train_split} | RS: {random_state}"
        ws.cell(row=row_of['Configuration'], column=new_col, value=config_desc)
        
        # Mark used features
        for p in combo_list:
            ws.cell(row=row_of[p], column=new_col, value='x')
            
        ws.cell(row=row_of['Output'], column=new_col, value=output)
        
        # Write metrics
        for m in METRIC_LABELS:
            ws.cell(row=row_of[m], column=new_col, value=round(metrics[m], 4) if not np.isnan(metrics[m]) else 'NaN')
            
        # Write feature importances
        for p in combo_list:
            if p in fi:
                ws.cell(row=row_of[f'FI: {p}'], column=new_col, value=round(float(fi[p]), 4))
                
        # Save dynamically at each step (safe but slower, strictly respects Approach 1)
        wb.save(filepath)

    print(f"\nExecution finished! Saved to {filepath}")

